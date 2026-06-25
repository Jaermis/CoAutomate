import openpyxl
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, OneCellAnchor

for fname in [
    'Certificate of Attendance CoA _ Kaquilala_Aug1_15.xlsx',
    'Certificate of Attendance CoA _ Kaquilala_Aug16_31.xlsx'
]:
    print("=" * 60)
    print("FILE:", fname)
    wb = openpyxl.load_workbook(fname)
    ws = wb.active
    print('Sheet name:', ws.title)
    print('Dimensions:', ws.dimensions)
    print('Max row:', ws.max_row, 'Max col:', ws.max_column)
    print()
    print('=== KEY CELLS ===')
    for cell_ref in ['B5','B6','B7','G8','A8','I37','J5','J6','A31']:
        cell = ws[cell_ref]
        print(cell_ref + ': value=' + repr(cell.value))
    print()
    print('=== IMAGES ===')
    for img in ws._images:
        anchor = img.anchor
        print('  Image anchor type:', type(anchor).__name__)
        try:
            if hasattr(anchor, '_from'):
                f = anchor._from
                print('  From cell: row=' + str(f.row) + ' col=' + str(f.col))
            if hasattr(anchor, 'to') and anchor.to:
                t = anchor.to
                print('  To cell: row=' + str(t.row) + ' col=' + str(t.col))
        except Exception as e:
            print('  Error reading anchor:', e)
        print('  Image format:', getattr(img, 'format', 'N/A'))
    print()
    print('=== MERGED CELLS (first 30) ===')
    for mc in list(ws.merged_cells.ranges)[:30]:
        print(' ', str(mc))
    print()
    print('=== ROW HEIGHTS (rows 1-44) ===')
    for r in range(1, 45):
        if r in ws.row_dimensions:
            rd = ws.row_dimensions[r]
            print(' Row', r, 'height:', rd.height)
    print()
    print('=== COL WIDTHS (A-L) ===')
    for c in ['A','B','C','D','E','F','G','H','I','J','K','L']:
        if c in ws.column_dimensions:
            cd = ws.column_dimensions[c]
            print(' Col', c, 'width:', cd.width)
